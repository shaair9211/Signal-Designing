import pygame
import sys
import random
import time
import math
import os
import openpyxl



class LeftTurn:
    def __init__(self):
        self.radius_lt = 16.5
        self.speed = 3
        self.dot_radii = 4.5
        self.stop_speed = 0
        self.reaction_time = 17
        self.spacing = 9
    def left_turn_i(self, point_x, point_y, angle, dot_x, dot_y, dot_x_nxt, dot_y_nxt):
        if dot_x < point_x and dot_y > point_y and phase_i_running == True:
            if dot_x+self.reaction_time >= dot_x_nxt:
                dot_x += self.stop_speed
            else:
                dot_x += self.speed
            pygame.draw.circle(screen, dark_cyan, (int(dot_x), int(dot_y)), self.dot_radii)
        elif dot_x < point_x and dot_y > point_y and phase_i_running == False:
            if point_x-10 < dot_x < point_x:
                dot_x += self.stop_speed
            elif dot_x+self.spacing >= dot_x_nxt:
                dot_x += self.stop_speed
            else:
                dot_x += self.speed
            pygame.draw.circle(screen, dark_cyan, (int(dot_x), int(dot_y)), self.dot_radii)
        elif dot_x >= point_x and dot_y > point_y and angle >= 0:
            dot_y = point_y + math.sin(angle) * self.radius_lt
            dot_x = point_x + math.cos(angle) * self.radius_lt
            angle -= 0.1
            pygame.draw.circle(screen, dark_cyan, (int(dot_x), int(dot_y)), self.dot_radii)
        else:
            dot_y -= self.speed
            pygame.draw.circle(screen, dark_cyan, (int(dot_x), int(dot_y)), self.dot_radii)
        return angle, dot_x, dot_y
    def left_turn_j(self, point_x, point_y, angle, dot_x, dot_y, dot_x_nxt, dot_y_nxt):
        if dot_y < point_y and dot_x < point_x and phase_j_running == True:
            if dot_y+self.reaction_time >= dot_y_nxt:
                dot_y += self.stop_speed
            else:
                dot_y += self.speed
            pygame.draw.circle(screen, dark_cyan, (int(dot_x), int(dot_y)), self.dot_radii)
        elif dot_y < point_y and dot_x < point_x and phase_j_running == False:
            if point_y-10 < dot_y < point_y:
                dot_y += self.stop_speed
            elif dot_y+self.spacing >= dot_y_nxt:
                dot_y += self.stop_speed
            else:
                dot_y += self.speed
            pygame.draw.circle(screen, dark_cyan, (int(dot_x), int(dot_y)), self.dot_radii)
        elif dot_y >= point_y and angle >= 0:
            dot_y = point_y + math.cos(angle) * self.radius_lt
            dot_x = point_x - math.sin(angle) * self.radius_lt
            angle -= 0.1
            pygame.draw.circle(screen, dark_cyan, (int(dot_x), int(dot_y)), self.dot_radii)
        else:
            dot_x += self.speed
            pygame.draw.circle(screen, dark_cyan, (int(dot_x), int(dot_y)), self.dot_radii)
        return angle, dot_x, dot_y
    def left_turn_k(self, point_x, point_y, angle, dot_x, dot_y, dot_x_nxt, dot_y_nxt):
        if dot_x > point_x and dot_y < point_y and phase_k_running == True:
            if dot_x-self.reaction_time <= dot_x_nxt:
                dot_x -= self.stop_speed
            else:
                dot_x -= self.speed
            pygame.draw.circle(screen, dark_cyan, (int(dot_x), int(dot_y)), self.dot_radii)
        elif dot_x > point_x and dot_y < point_y and phase_k_running == False:
            if point_x+10 > dot_x > point_x:
                dot_x -= self.stop_speed
            elif dot_x-self.spacing <= dot_x_nxt:
                dot_x -= self.stop_speed
            else:
                dot_x -= self.speed
            pygame.draw.circle(screen, dark_cyan, (int(dot_x), int(dot_y)), self.dot_radii)
        elif dot_x <= point_x and angle >= 0:
            dot_y = point_y - math.sin(angle) * self.radius_lt
            dot_x = point_x - math.cos(angle) * self.radius_lt
            angle -= 0.1
            pygame.draw.circle(screen, dark_cyan, (int(dot_x), int(dot_y)), self.dot_radii)
        else:
            dot_y += self.speed
            pygame.draw.circle(screen, dark_cyan, (int(dot_x), int(dot_y)), self.dot_radii)
        return angle, dot_x, dot_y
    def left_turn_l(self, point_x, point_y, angle, dot_x, dot_y, dot_x_nxt, dot_y_nxt):
        if dot_y > point_y and dot_x > point_x and phase_l_running == True:
            if dot_y-self.reaction_time <= dot_y_nxt:
                dot_y -= self.stop_speed
            else:
                dot_y -= self.speed
            pygame.draw.circle(screen, dark_cyan, (int(dot_x), int(dot_y)), self.dot_radii)
        elif dot_y > point_y and dot_x > point_x and phase_l_running == False:
            if point_y+11 > dot_y > point_y:
                dot_y -= self.stop_speed
            elif dot_y-self.spacing <= dot_y_nxt:
                dot_y -= self.stop_speed
            else:
                dot_y -= self.speed
            pygame.draw.circle(screen, dark_cyan, (int(dot_x), int(dot_y)), self.dot_radii)
        elif dot_y <= point_y and angle >= 0:
            dot_y = point_y - math.cos(angle) * self.radius_lt
            dot_x = point_x + math.sin(angle) * self.radius_lt
            angle -= 0.1
            pygame.draw.circle(screen, dark_cyan, (int(dot_x), int(dot_y)), self.dot_radii)
        else:
            dot_x -= self.speed
            pygame.draw.circle(screen, dark_cyan, (int(dot_x), int(dot_y)), self.dot_radii)
        return angle, dot_x, dot_y
left_turn = LeftTurn()

class Thru:
    def __init__(self):
        self.speed = 3
        self.dot_radii = 4.5
        self.stop_speed = 0
        self.reaction_time = 30
        self.spacing = 9
    def thru_i(self, dot_x, dot_y, dot_x_nxt, dot_y_nxt):
        if dot_x <= screen_width and phase_i_running == True:
            if dot_x+self.reaction_time >= dot_x_nxt and dot_x_nxt < screen_width:
                dot_x += self.stop_speed
            else:
                dot_x += self.speed
            pygame.draw.circle(screen, light_red, (int(dot_x), int(dot_y)), self.dot_radii)
        elif dot_x <= screen_width and phase_i_running == False:
            if 501-10 < dot_x < 501:
                dot_x += self.stop_speed
            elif dot_x+self.spacing >= dot_x_nxt and dot_x_nxt < screen_width:
                dot_x += self.stop_speed
            else:
                dot_x += self.speed
            pygame.draw.circle(screen, light_red, (int(dot_x), int(dot_y)), self.dot_radii)
        return dot_x, dot_y
    def thru_j(self, dot_x, dot_y, dot_x_nxt, dot_y_nxt):
        if dot_y <= screen_height and phase_j_running == True:
            if dot_y+self.reaction_time >= dot_y_nxt and dot_y_nxt < screen_height:
                dot_y += self.stop_speed
            else:
                dot_y += self.speed
            pygame.draw.circle(screen, light_red, (int(dot_x), int(dot_y)), self.dot_radii)
        elif dot_y <= screen_height and phase_j_running == False:
            if 301-10 < dot_y < 301:
                dot_y += self.stop_speed
            elif dot_y+self.spacing >= dot_y_nxt and dot_y_nxt < screen_height:
                dot_x += self.stop_speed
            else:
                dot_y += self.speed
            pygame.draw.circle(screen, light_red, (int(dot_x), int(dot_y)), self.dot_radii)
        return dot_x, dot_y
    def thru_k(self, dot_x, dot_y, dot_x_nxt, dot_y_nxt):
        if dot_x >= 0 and phase_k_running == True:
            if dot_x-self.reaction_time <= dot_x_nxt and dot_x_nxt > 0:
                dot_x -= self.stop_speed
            else:
                dot_x -= self.speed
            pygame.draw.circle(screen, light_red, (int(dot_x), int(dot_y)), self.dot_radii)
        elif dot_x >= 0 and phase_k_running == False:
            if 699+10 > dot_x > 699:
                dot_x -= self.stop_speed
            elif dot_x-self.spacing <= dot_x_nxt and dot_x_nxt > 0:
                dot_x -= self.stop_speed
            else:
                dot_x -= self.speed
            pygame.draw.circle(screen, light_red, (int(dot_x), int(dot_y)), self.dot_radii)
        return dot_x, dot_y
    def thru_l(self, dot_x, dot_y, dot_x_nxt, dot_y_nxt):
        if dot_y >= 0 and phase_l_running == True:
            if dot_y-self.reaction_time <= dot_y_nxt and dot_y_nxt > 0:
                dot_y -= self.stop_speed
            else:
                dot_y -= self.speed
            pygame.draw.circle(screen, light_red, (int(dot_x), int(dot_y)), self.dot_radii)
        elif dot_y >= 0 and phase_l_running == False:
            if 499+12 > dot_y > 499:
                dot_y -= self.stop_speed
            elif dot_y-self.spacing <= dot_y_nxt and dot_y_nxt > 0:
                dot_y -= self.stop_speed
            else:
                dot_y -= self.speed
            pygame.draw.circle(screen, light_red, (int(dot_x), int(dot_y)), self.dot_radii)
        return dot_x, dot_y
thru = Thru()

class RightTurn:
    def __init__(self):
        self.radius_lt = 115.5
        self.speed = 3
        self.dot_radii = 4.5
        self.stop_speed = 0
        self.reaction_time = 30
        self.spacing = 9
    def right_turn_i(self, point_x, point_y, angle, dot_x, dot_y, dot_x_nxt, dot_y_nxt):
        if dot_x < point_x and dot_y < point_y and phase_i_running == True:
            if dot_x+self.reaction_time >= dot_x_nxt:
                dot_x += self.stop_speed
            else:
                dot_x += self.speed
            pygame.draw.circle(screen, purple, (int(dot_x), int(dot_y)), self.dot_radii)
        elif dot_x < point_x and dot_y < point_y and phase_i_running == False:
            if point_x-10 < dot_x < point_x:
                dot_x += self.stop_speed
            elif dot_x+self.spacing >= dot_x_nxt:
                dot_x += self.stop_speed
            else:
                dot_x += self.speed
            pygame.draw.circle(screen, purple, (int(dot_x), int(dot_y)), self.dot_radii)
        elif dot_x >= point_x and angle >= 0:
            dot_y = point_y - math.sin(angle) * self.radius_lt
            dot_x = point_x + math.cos(angle) * self.radius_lt
            angle -= 0.03
            pygame.draw.circle(screen, purple, (int(dot_x), int(dot_y)), self.dot_radii)
        else:
            dot_y += self.speed
            pygame.draw.circle(screen, purple, (int(dot_x), int(dot_y)), self.dot_radii)
        return angle, dot_x, dot_y
    def right_turn_j(self, point_x, point_y, angle, dot_x, dot_y, dot_x_nxt, dot_y_nxt):
        if dot_y < point_y and dot_x > point_x and phase_j_running == True:
            if dot_y+self.reaction_time >= dot_y_nxt:
                dot_y += self.stop_speed
            else:
                dot_y += self.speed
            pygame.draw.circle(screen, purple, (int(dot_x), int(dot_y)), self.dot_radii)
        elif dot_y < point_y and dot_x > point_x and phase_j_running == False:
            if point_y-10 < dot_y < point_y:
                dot_y += self.stop_speed
            elif dot_y+self.spacing >= dot_y_nxt:
                dot_x += self.stop_speed
            else:
                dot_y += self.speed
            pygame.draw.circle(screen, purple, (int(dot_x), int(dot_y)), self.dot_radii)
        elif dot_y >= point_y and angle >= 0:
            dot_y = point_y + math.cos(angle) * self.radius_lt
            dot_x = point_x + math.sin(angle) * self.radius_lt
            angle -= 0.03
            pygame.draw.circle(screen, purple, (int(dot_x), int(dot_y)), self.dot_radii)
        else:
            dot_x -= self.speed
            pygame.draw.circle(screen, purple, (int(dot_x), int(dot_y)), self.dot_radii)
        return angle, dot_x, dot_y
    def right_turn_k(self, point_x, point_y, angle, dot_x, dot_y, dot_x_nxt, dot_y_nxt):
        if dot_x > point_x and dot_y > point_y and phase_k_running == True:
            if dot_x-self.reaction_time <= dot_x_nxt:
                dot_x -= self.stop_speed
            else:
                dot_x -= self.speed
            pygame.draw.circle(screen, purple, (int(dot_x), int(dot_y)), self.dot_radii)
        elif dot_x > point_x and dot_y > point_y and phase_k_running == False:
            if point_x+10 > dot_x > point_x:
                dot_x -= self.stop_speed
            elif dot_x-self.spacing <= dot_x_nxt:
                dot_x -= self.stop_speed
            else:
                dot_x -= self.speed
            pygame.draw.circle(screen, purple, (int(dot_x), int(dot_y)), self.dot_radii)
        elif dot_x <= point_x and angle >= 0:
            dot_y = point_y + math.sin(angle) * self.radius_lt
            dot_x = point_x - math.cos(angle) * self.radius_lt
            angle -= 0.03
            pygame.draw.circle(screen, purple, (int(dot_x), int(dot_y)), self.dot_radii)
        else:
            dot_y -= self.speed
            pygame.draw.circle(screen, purple, (int(dot_x), int(dot_y)), self.dot_radii)
        return angle, dot_x, dot_y
    def right_turn_l(self, point_x, point_y, angle, dot_x, dot_y, dot_x_nxt, dot_y_nxt):
        if dot_y > point_y and dot_x < point_x and phase_l_running == True:
            if dot_y-self.reaction_time <= dot_y_nxt:
                dot_y -= self.stop_speed
            else:
                dot_y -= self.speed
            pygame.draw.circle(screen, purple, (int(dot_x), int(dot_y)), self.dot_radii)
        elif dot_y > point_y and dot_x < point_x and phase_l_running == False:
            if point_y+11 > dot_y > point_y:
                dot_y -= self.stop_speed
            elif dot_y-self.spacing <= dot_y_nxt:
                dot_y -= self.stop_speed
            else:
                dot_y -= self.speed
            pygame.draw.circle(screen, purple, (int(dot_x), int(dot_y)), self.dot_radii)
        elif dot_y <= point_y and angle >= 0:
            dot_y = point_y - math.cos(angle) * self.radius_lt
            dot_x = point_x - math.sin(angle) * self.radius_lt
            angle -= 0.03
            pygame.draw.circle(screen, purple, (int(dot_x), int(dot_y)), self.dot_radii)
        else:
            dot_x += self.speed
            pygame.draw.circle(screen, purple, (int(dot_x), int(dot_y)), self.dot_radii)
        return angle, dot_x, dot_y
right_turn = RightTurn()





class RedLight:
    def __init__(self):
        self.dot_radii = 8
    def red_light_i():
        pygame.draw.circle(screen, red_light, (431, 269.5), 10)
    def red_light_j():
        pygame.draw.circle(screen, red_light, (739.5, 231), 10)
    def red_light_k():
        pygame.draw.circle(screen, red_light, (769, 539.5), 10)
    def red_light_l():
        pygame.draw.circle(screen, red_light, (469.5, 569), 10)

class YellowLight:
    def __init__(self):
        self.dot_radii = 8
    def yellow_light_i():
        pygame.draw.circle(screen, yellow_light, (431, 269.5), 10)
    def yellow_light_j():
        pygame.draw.circle(screen, yellow_light, (739.5, 231), 10)
    def yellow_light_k():
        pygame.draw.circle(screen, yellow_light, (769, 539.5), 10)
    def yellow_light_l():
        pygame.draw.circle(screen, yellow_light, (469.5, 569), 10)

class GreenLight:
    def __init__(self):
        self.dot_radii = 8
    def green_light_i():
        pygame.draw.circle(screen, green_light, (431, 269.5), 10)
    def green_light_j():
        pygame.draw.circle(screen, green_light, (739.5, 231), 10)
    def green_light_k():
        pygame.draw.circle(screen, green_light, (769, 539.5), 10)
    def green_light_l():
        pygame.draw.circle(screen, green_light, (469.5, 569), 10)


i = 0
j = 0
k = 0
l = 0

def phase_i():
    global phase_i_running, phase_j_running, phase_k_running, phase_l_running
    phase_j_running, phase_k_running, phase_l_running = False, False, False
    global i,j,k,l
    j,k,l = 0, 0, 0
    if i <= 180:
        YellowLight.yellow_light_i()
        i+=1
    else:
        GreenLight.green_light_i()
    RedLight.red_light_j()
    RedLight.red_light_k()
    RedLight.red_light_l()
    phase_i_running = True

def phase_j():
    global phase_i_running, phase_j_running, phase_k_running, phase_l_running
    phase_i_running, phase_k_running, phase_l_running = False, False, False
    global i,j,k,l
    i,k,l = 0, 0, 0
    if j <= 180:
        YellowLight.yellow_light_j()
        j+=1
    else:
        GreenLight.green_light_j()
    RedLight.red_light_i()
    RedLight.red_light_k()
    RedLight.red_light_l()
    phase_j_running = True

def phase_k():
    global phase_i_running, phase_j_running, phase_k_running, phase_l_running
    phase_i_running, phase_j_running, phase_l_running = False, False, False
    global i,j,k,l
    i,j,l = 0, 0, 0
    if k <= 180:
        YellowLight.yellow_light_k()
        k+=1
    else:
        GreenLight.green_light_k()
    RedLight.red_light_i()
    RedLight.red_light_j()
    RedLight.red_light_l()
    phase_k_running = True

def phase_l():
    global phase_i_running, phase_j_running, phase_k_running, phase_l_running
    phase_i_running, phase_j_running, phase_k_running = False, False, False
    global i,j,k,l
    i,j,k = 0, 0, 0
    if l <= 180:
        YellowLight.yellow_light_l()
        l+=1
    else:
        GreenLight.green_light_l()
    RedLight.red_light_i()
    RedLight.red_light_j()
    RedLight.red_light_k()
    phase_l_running = True













pygame.init()
screen_width, screen_height = 1200, 800
screen = pygame.display.set_mode((screen_width, screen_height))
pygame.display.set_caption("Traffic Intersection Simulation")
white, black, asphalt_grey, green, yellow, dark_cyan, purple, light_red = (255, 255, 255), (0, 0, 0), (99, 102, 102), (1, 50, 32), (255, 255, 0), (14, 234, 255), (149, 117, 205), (255, 82, 82)
red_light, green_light, yellow_light = (220,20,60), (124,252,0), (255,255,0)





dot_lix, dot_liy = 0, 317.5
dot_ljx, dot_ljy = 682.5, 0
dot_lkx, dot_lky = 1200, 482.5
dot_llx, dot_lly = 517.5, 800
angle = 1.57
dot_dict_li = {}
dot_dict_lj = {}
dot_dict_lk = {}
dot_dict_ll = {}
index_li = 0
index_lj = 0
index_lk = 0
index_ll = 0
dot_tix, dot_tiy = 0, 350.5
dot_tjx, dot_tjy = 649.5, 0
dot_tkx, dot_tky = 1200, 449.5
dot_tlx, dot_tly = 549.5, 800
dot_dict_ti = {}
dot_dict_tj = {}
dot_dict_tk = {}
dot_dict_tl = {}
index_ti = 0
index_tj = 0
index_tk = 0
index_tl = 0
dot_rix, dot_riy = 0, 383.5
dot_rjx, dot_rjy = 616.5, 0
dot_rkx, dot_rky = 1200, 416.5
dot_rlx, dot_rly = 583.5, 800
angle = 1.57
dot_dict_ri = {}
dot_dict_rj = {}
dot_dict_rk = {}
dot_dict_rl = {}
index_ri = 0
index_rj = 0
index_rk = 0
index_rl = 0
density_i = []
density_j = []
density_k = []
density_l = []


shararti_molana = 0
karan_johar = 0
fps = 60
frame_count = 0
timer = -1
t = True
running = True
clock = pygame.time.Clock()
while shararti_molana < 150*fps:
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            running = False
    screen.fill(asphalt_grey)
        #defining corner rectangles
    rect_width, rect_height, rect_margin = 500,300, 0
            # Draw rectangles on each corner of the window
        # Top-left corner
    pygame.draw.rect(screen, green, (rect_margin, rect_margin, rect_width + 1, rect_height + 1))
        # Top-right corner
    pygame.draw.rect(screen, green, (screen_width - rect_margin - rect_width + 1, rect_margin, rect_width, rect_height + 1))
        # Bottom-left corner
    pygame.draw.rect(screen, green, (rect_margin + 1, screen_height - rect_margin - rect_height + 1, rect_width, rect_height))
        # Bottom-right corner
    pygame.draw.rect(screen, green, (screen_width - rect_margin - rect_width + 1, screen_height - rect_margin - rect_height + 1, rect_width, rect_height))
        #dotted line properties
    dashed_line_width = 2
    dashed_line_length_h = 500
    dashed_line_length_v = 300
    dash_gap = 40
    dash_length = 20
        #lane specification
    lane_width = 33        
        # Bottom portion of vertical plain lines
    p = screen_height - dashed_line_length_v
    horiz_center = screen_width // 2
    pygame.draw.line(screen, white, ((horiz_center), p), ((screen_width // 2), (p + dashed_line_length_v)), dashed_line_width)
    pygame.draw.line(screen, black, ((horiz_center - (3*lane_width)), p), ((horiz_center - (3*lane_width)), (p + dashed_line_length_v)),    dashed_line_width)    
    pygame.draw.line(screen, black, ((horiz_center + ((3*lane_width) )), p), ((horiz_center + (3*lane_width)), (p + dashed_line_length_v)),    dashed_line_width)    
    pygame.draw.line(screen, white, ((horiz_center - ((3*lane_width + 2) )), (p - 199)), ((horiz_center - (3*lane_width + 2)), (p - 100)),    dashed_line_width + 3)    
    pygame.draw.line(screen, white, ((horiz_center + ((3*lane_width + 2) )), (p - 100)), ((horiz_center + (3*lane_width + 2)), (p - 1)),    dashed_line_width + 3)    
        # Bottom portion of vertical dashed lines
    for y in range(p, screen_height, dash_gap):
        pygame.draw.line(screen, white, ((horiz_center  - (lane_width)), y), ((horiz_center - (lane_width)), y + dash_length),   dashed_line_width)
        pygame.draw.line(screen, white, ((horiz_center  - (2*lane_width)), y), ((horiz_center - (2*lane_width)), y + dash_length),   dashed_line_width)
        pygame.draw.line(screen, yellow, ((horiz_center  - (3*lane_width)), y -1), ((horiz_center - (3*lane_width)), y + dash_length),  dashed_line_width)
        pygame.draw.line(screen, white, ((horiz_center  + (lane_width)), y), ((horiz_center + (lane_width)), y + dash_length),   dashed_line_width)
        pygame.draw.line(screen, white, ((horiz_center  + (2*lane_width)), y), ((horiz_center + (2*lane_width)), y + dash_length),   dashed_line_width)
        pygame.draw.line(screen, yellow, ((horiz_center  + ((3*lane_width) )), y -1), ((horiz_center + (3*lane_width)), y + dash_length),  dashed_line_width)
        # Top portion of vertical plain lines
    r = dashed_line_length_v
    horiz_center = screen_width // 2
    pygame.draw.line(screen, white, ((horiz_center), 0), ((screen_width // 2), (0 + dashed_line_length_v)), dashed_line_width)
    pygame.draw.line(screen, black, ((horiz_center - (3*lane_width)), 0), ((horiz_center - (3*lane_width)), (0 + dashed_line_length_v)),    dashed_line_width)    
    pygame.draw.line(screen, black, ((horiz_center + ((3*lane_width) )), 0), ((horiz_center + (3*lane_width)), (0 + dashed_line_length_v)),    dashed_line_width)    
        # Top portion of vertical dashed lines
    for y in range(0, r, dash_gap):
        pygame.draw.line(screen, white, ((horiz_center  - (lane_width)), y), ((horiz_center - (lane_width)), y + dash_length),   dashed_line_width)
        pygame.draw.line(screen, white, ((horiz_center  - (2*lane_width)), y), ((horiz_center - (2*lane_width)), y + dash_length),   dashed_line_width)
        pygame.draw.line(screen, yellow, ((horiz_center  - (3*lane_width)), y), ((horiz_center - (3*lane_width)), y + dash_length +2),  dashed_line_width)
        pygame.draw.line(screen, white, ((horiz_center  + (lane_width)), y), ((horiz_center + (lane_width)), y + dash_length),   dashed_line_width)
        pygame.draw.line(screen, white, ((horiz_center  + (2*lane_width)), y), ((horiz_center + (2*lane_width)), y + dash_length),   dashed_line_width)
        pygame.draw.line(screen, yellow, ((horiz_center  + ((3*lane_width) )), y), ((horiz_center + (3*lane_width)), y + dash_length +2),  dashed_line_width)
        # Right portion of horizontal plain lines
    q = screen_width - dashed_line_length_h
    vert_center = screen_height // 2
    pygame.draw.line(screen, white, (q, (vert_center)), ((q + dashed_line_length_h), (vert_center)),    dashed_line_width)
    pygame.draw.line(screen, black, (q, (vert_center - (3*lane_width))), ((q + dashed_line_length_h), (vert_center - (3*lane_width))),  dashed_line_width)
    pygame.draw.line(screen, black, (q, (vert_center + (3*lane_width))), ((q + dashed_line_length_h), (vert_center + (3*lane_width))),  dashed_line_width)
    pygame.draw.line(screen, white, ((q - 100), (vert_center - (3*lane_width + 1))), ((q - 2), (vert_center - (3*lane_width + 1))),  dashed_line_width +3)
    pygame.draw.line(screen, white, ((q - 198), (vert_center + (3*lane_width + 2))), ((q - 100), (vert_center + (3*lane_width + 2))),  dashed_line_width +3)
        # Right portion of horizontal dashed lines
    for x in range(q, screen_width, dash_gap):
        pygame.draw.line(screen, white, (x, (vert_center - (lane_width))), (x + dash_length, (vert_center - (lane_width))),    dashed_line_width)
        pygame.draw.line(screen, white, (x, (vert_center - (2*lane_width))), (x + dash_length, (vert_center - (2*lane_width))),    dashed_line_width)
        pygame.draw.line(screen, yellow, (x, (vert_center - (3*lane_width))), (x + dash_length, (vert_center - (3*lane_width))),   dashed_line_width)
        pygame.draw.line(screen, white, (x, (vert_center + (lane_width))), (x + dash_length, (vert_center + (lane_width))),    dashed_line_width)
        pygame.draw.line(screen, white, (x, (vert_center + (2*lane_width))), (x + dash_length, (vert_center + (2*lane_width))),    dashed_line_width)
        pygame.draw.line(screen, yellow, (x, (vert_center + (3*lane_width))), (x + dash_length, (vert_center + (3*lane_width))),   dashed_line_width)
        # Left portion of horizontal plain lines
    s = dashed_line_length_h
    vert_center = screen_height // 2
    pygame.draw.line(screen, white, (0, (vert_center)), ((0 + dashed_line_length_h), (vert_center)),    dashed_line_width)
    pygame.draw.line(screen, black, (0, (vert_center + (3*lane_width))), ((0 + dashed_line_length_h), (vert_center + (3*lane_width))),  dashed_line_width)
    pygame.draw.line(screen, black, (0, (vert_center - (3*lane_width))), ((0 + dashed_line_length_h), (vert_center - (3*lane_width))),  dashed_line_width)
        # Left portion of horizontal dashed lines
    for x in range(0, s, dash_gap):
        pygame.draw.line(screen, white, (x, (vert_center - (lane_width))), (x + dash_length, (vert_center - (lane_width))),    dashed_line_width)
        pygame.draw.line(screen, white, (x, (vert_center - (2*lane_width))), (x + dash_length, (vert_center - (2*lane_width))),    dashed_line_width)
        pygame.draw.line(screen, yellow, (x, (vert_center - (3*lane_width))), (x + dash_length, (vert_center - (3*lane_width))),   dashed_line_width)
        pygame.draw.line(screen, white, (x, (vert_center + (lane_width))), (x + dash_length, (vert_center + (lane_width))),    dashed_line_width)
        pygame.draw.line(screen, white, (x, (vert_center + (2*lane_width))), (x + dash_length, (vert_center + (2*lane_width))),    dashed_line_width)
        pygame.draw.line(screen, yellow, (x, (vert_center + (3*lane_width))), (x + dash_length, (vert_center + (3*lane_width))),   dashed_line_width)







    dots_to_remove_li = []
    dots_to_remove_lj = []
    dots_to_remove_lk = []
    dots_to_remove_ll = []
    dots_to_remove_ti = []
    dots_to_remove_tj = []
    dots_to_remove_tk = []
    dots_to_remove_tl = []
    dots_to_remove_ri = []
    dots_to_remove_rj = []
    dots_to_remove_rk = []
    dots_to_remove_rl = []
    expectancy = 0.03
    expectancy_i, expectancy_j, expectancy_k, expectancy_l = (expectancy*0.1108), (expectancy*0.3612), (expectancy*0.1047), (expectancy*0.4232)
    len_li = 0
    len_ti = 0
    len_ri = 0
    len_lj = 0
    len_tj = 0
    len_rj = 0
    len_lk = 0
    len_tk = 0
    len_rk = 0
    len_ll = 0
    len_tl = 0
    len_rl = 0



    if random.random() < expectancy_i:
        dot_dict_li[index_li] = (501, 301, 1.57, 0, 317.5)
        index_li += 1
    for dot_index in dot_dict_li:
        point_x, point_y, angle, dot_lix, dot_liy = dot_dict_li[dot_index]
        if dot_index-1 in dot_dict_li:
            point_x_nxt, point_y_nxt, angle_nxt, dot_lix_nxt, dot_liy_nxt = dot_dict_li[dot_index - 1]
        else:
            dot_lix_nxt, dot_liy_nxt = 1200, 800
        angle, dot_lix, dot_liy = left_turn.left_turn_i(point_x, point_y, angle, dot_lix, dot_liy, dot_lix_nxt, dot_liy_nxt)
        dot_dict_li[dot_index] = (point_x, point_y, angle, dot_lix, dot_liy)
        x_cord_li = dot_dict_li[dot_index][3]
        if x_cord_li < 501:
            len_li +=1
        if dot_lix < 0 or dot_liy < 0 or dot_lix > screen_width or dot_liy > screen_height:
            dots_to_remove_li.append(dot_index)
    for index in dots_to_remove_li:
        del dot_dict_li[index]
    if random.random() < expectancy_j:
        dot_dict_lj[index_lj] = (699, 301, 1.57, 682.5, 0)
        index_lj += 1
    for dot_index in dot_dict_lj:
        point_x, point_y, angle, dot_ljx, dot_ljy = dot_dict_lj[dot_index]
        if dot_index-1 in dot_dict_lj:
            point_x_nxt, point_y_nxt, angle_nxt, dot_ljx_nxt, dot_ljy_nxt = dot_dict_lj[dot_index - 1]
        else:
            dot_ljx_nxt, dot_ljy_nxt = 1200, 800
        angle, dot_ljx, dot_ljy = left_turn.left_turn_j(point_x, point_y, angle, dot_ljx, dot_ljy, dot_ljx_nxt, dot_ljy_nxt)
        dot_dict_lj[dot_index] = (point_x, point_y, angle, dot_ljx, dot_ljy)
        y_cord_lj = dot_dict_lj[dot_index][4]
        if y_cord_lj < 301:
            len_lj +=1
        if dot_ljx < 0 or dot_ljy < 0 or dot_ljx > screen_width or dot_lly > screen_height:
            dots_to_remove_lj.append(dot_index)
    for index in dots_to_remove_lj:
        del dot_dict_lj[index]
    if random.random() < expectancy_k:
        dot_dict_lk[index_lk] = (699, 499, 1.57, 1200, 482.5)
        index_lk += 1
    for dot_index in dot_dict_lk:
        point_x, point_y, angle, dot_lkx, dot_lky = dot_dict_lk[dot_index]
        if dot_index-1 in dot_dict_lk:
            point_x_nxt, point_y_nxt, angle_nxt, dot_lkx_nxt, dot_lky_nxt = dot_dict_lk[dot_index - 1]
        else:
            dot_lkx_nxt, dot_lky_nxt = 0, 0
        angle, dot_lkx, dot_lky = left_turn.left_turn_k(point_x, point_y, angle, dot_lkx, dot_lky, dot_lkx_nxt, dot_lky_nxt)
        dot_dict_lk[dot_index] = (point_x, point_y, angle, dot_lkx, dot_lky)
        x_cord_lk = dot_dict_lk[dot_index][3]
        if x_cord_lk > 699:
            len_lk +=1
        if dot_lkx < 0 or dot_lky < 0 or dot_lkx > screen_width or dot_lky > screen_height:
            dots_to_remove_lk.append(dot_index)
    for index in dots_to_remove_lk:
        del dot_dict_lk[index]
    if random.random() < expectancy_l:
        dot_dict_ll[index_ll] = (501, 499, 1.57, 517.5, 800)
        index_ll += 1
    for dot_index in dot_dict_ll:
        point_x, point_y, angle, dot_llx, dot_lly = dot_dict_ll[dot_index]
        if dot_index-1 in dot_dict_ll:
            point_x_nxt, point_y_nxt, angle_nxt, dot_llx_nxt, dot_lly_nxt = dot_dict_ll[dot_index - 1]
        else:
            dot_llx_nxt, dot_lly_nxt = 0, 0
        angle, dot_llx, dot_lly = left_turn.left_turn_l(point_x, point_y, angle, dot_llx, dot_lly, dot_llx_nxt, dot_lly_nxt)
        dot_dict_ll[dot_index] = (point_x, point_y, angle, dot_llx, dot_lly)
        y_cord_ll = dot_dict_ll[dot_index][4]
        if y_cord_ll > 499:
            len_ll +=1
        if dot_llx < 0 or dot_lly < 0 or dot_llx > screen_width or dot_lly > screen_height:
            dots_to_remove_ll.append(dot_index)
    for index in dots_to_remove_ll:
        del dot_dict_ll[index]



    if random.random() < expectancy_i:
        dot_dict_ti[index_ti] = (0, 350.5)
        index_ti += 1
    for dot_index in dot_dict_ti:
        dot_tix, dot_tiy = dot_dict_ti[dot_index]
        if dot_index-1 in dot_dict_ti:
            dot_tix_nxt, dot_tiy_nxt = dot_dict_ti[dot_index - 1]
        else:
            dot_tix_nxt, dot_tiy_nxt = 1200, 800
        dot_tix, dot_tiy = thru.thru_i(dot_tix, dot_tiy, dot_tix_nxt, dot_tiy_nxt)
        dot_dict_ti[dot_index] = (dot_tix, dot_tiy)
        x_cord_ti = dot_dict_ti[dot_index][0]
        if x_cord_ti < 501:
            len_ti += 1
        if dot_tix < 0 or dot_tiy < 0 or dot_tix > screen_width or dot_tiy > screen_height:
            dots_to_remove_ti.append(dot_index)
    for index in dots_to_remove_ti:
        del dot_dict_ti[index]
    if random.random() < expectancy_j:
        dot_dict_tj[index_tj] = (649.5, 0)
        index_tj += 1
    for dot_index in dot_dict_tj:
        dot_tjx, dot_tjy = dot_dict_tj[dot_index]
        if dot_index-1 in dot_dict_tj:
            dot_tjx_nxt, dot_tjy_nxt = dot_dict_tj[dot_index - 1]
        else:
            dot_tjx_nxt, dot_tjy_nxt = 1200, 800
        dot_tjx, dot_tjy = thru.thru_j(dot_tjx, dot_tjy, dot_tjx_nxt, dot_tjy_nxt)
        dot_dict_tj[dot_index] = (dot_tjx, dot_tjy)
        y_cord_tj = dot_dict_tj[dot_index][1]
        if y_cord_tj < 301:
            len_tj +=1
        if dot_tjx < 0 or dot_tjy < 0 or dot_tjx >= screen_width or dot_tly >= screen_height:
            dots_to_remove_tj.append(dot_index)
    for index in dots_to_remove_tj:
        del dot_dict_tj[index]
    if random.random() < expectancy_k:
        dot_dict_tk[index_tk] = (1200, 449.5)
        index_tk += 1
    for dot_index in dot_dict_tk:
        dot_tkx, dot_tky = dot_dict_tk[dot_index]
        if dot_index-1 in dot_dict_tk:
            dot_tkx_nxt, dot_tky_nxt = dot_dict_tk[dot_index - 1]
        else:
            dot_tkx_nxt, dot_tky_nxt = 0, 0
        dot_tkx, dot_tky = thru.thru_k(dot_tkx, dot_tky, dot_tkx_nxt, dot_tky_nxt)
        dot_dict_tk[dot_index] = (dot_tkx, dot_tky)
        x_cord_tk = dot_dict_tk[dot_index][0]
        if x_cord_tk > 699:
            len_tk +=1
        if dot_tkx < 0 or dot_tky < 0 or dot_tkx > screen_width or dot_tky > screen_height:
            dots_to_remove_tk.append(dot_index)
    for index in dots_to_remove_tk:
        del dot_dict_tk[index]
    if random.random() < expectancy_l:
        dot_dict_tl[index_tl] = (549.5, 800)
        index_tl += 1
    for dot_index in dot_dict_tl:
        dot_tlx, dot_tly = dot_dict_tl[dot_index]
        if dot_index-1 in dot_dict_tl:
            dot_tlx_nxt, dot_tly_nxt = dot_dict_tl[dot_index - 1]
        else:
            dot_tlx_nxt, dot_tly_nxt = 0, 0
        dot_tlx, dot_tly = thru.thru_l(dot_tlx, dot_tly, dot_tlx_nxt, dot_tly_nxt)
        dot_dict_tl[dot_index] = (dot_tlx, dot_tly)
        y_cord_tl = dot_dict_tl[dot_index][1]
        if y_cord_tl > 499:
            len_tl +=1
        if dot_tlx < 0 or dot_tly < 0 or dot_tlx > screen_width or dot_tly > screen_height:
            dots_to_remove_tl.append(dot_index)
    for index in dots_to_remove_tl:
        del dot_dict_tl[index]



    if random.random() < expectancy_i:
        dot_dict_ri[index_ri] = (501, 499, 1.57, 0, 383.5)
        index_ri += 1
    for dot_index in dot_dict_ri:
        point_x, point_y, angle, dot_rix, dot_riy = dot_dict_ri[dot_index]
        if dot_index-1 in dot_dict_ri:
            point_x_nxt, point_y_nxt, angle_nxt, dot_rix_nxt, dot_riy_nxt = dot_dict_ri[dot_index - 1]
        else:
            dot_rix_nxt, dot_riy_nxt = 1200, 800
        angle, dot_rix, dot_riy = right_turn.right_turn_i(point_x, point_y, angle, dot_rix, dot_riy, dot_rix_nxt, dot_riy_nxt)
        dot_dict_ri[dot_index] = (point_x, point_y, angle, dot_rix, dot_riy)
        x_cord_ri = dot_dict_ri[dot_index][3]
        if x_cord_ri < 501:
            len_ri += 1
        if dot_rix < 0 or dot_riy < 0 or dot_rix > screen_width or dot_riy > screen_height:
            dots_to_remove_ri.append(dot_index)
    for index in dots_to_remove_ri:
        del dot_dict_ri[index]
    if random.random() < expectancy_j:
        dot_dict_rj[index_rj] = (501, 301, 1.57, 616.5, 0)
        index_rj += 1
    for dot_index in dot_dict_rj:
        point_x, point_y, angle, dot_rjx, dot_rjy = dot_dict_rj[dot_index]
        if dot_index-1 in dot_dict_rj:
            point_x_nxt, point_y_nxt, angle_nxt, dot_rjx_nxt, dot_rjy_nxt = dot_dict_rj[dot_index - 1]
        else:
            dot_rjx_nxt, dot_rjy_nxt = 1200, 800
        angle, dot_rjx, dot_rjy = right_turn.right_turn_j(point_x, point_y, angle, dot_rjx, dot_rjy, dot_rjx_nxt, dot_rjy_nxt)
        dot_dict_rj[dot_index] = (point_x, point_y, angle, dot_rjx, dot_rjy)
        y_cord_rj = dot_dict_rj[dot_index][4]
        if y_cord_rj < 301:
            len_rj +=1
        if dot_rjx < 0 or dot_rjy < 0 or dot_rjx > screen_width or dot_rly > screen_height:
            dots_to_remove_rj.append(dot_index)
    for index in dots_to_remove_rj:
        del dot_dict_rj[index]
    if random.random() < expectancy_k:
        dot_dict_rk[index_rk] = (699, 301, 1.57, 1200, 416.5)
        index_rk += 1
    for dot_index in dot_dict_rk:
        point_x, point_y, angle, dot_rkx, dot_rky = dot_dict_rk[dot_index]
        if dot_index-1 in dot_dict_rk:
            point_x_nxt, point_y_nxt, angle_nxt, dot_rkx_nxt, dot_rky_nxt = dot_dict_rk[dot_index - 1]
        else:
            dot_rkx_nxt, dot_rky_nxt = 0, 0
        angle, dot_rkx, dot_rky = right_turn.right_turn_k(point_x, point_y, angle, dot_rkx, dot_rky, dot_rkx_nxt, dot_rky_nxt)
        dot_dict_rk[dot_index] = (point_x, point_y, angle, dot_rkx, dot_rky)
        x_cord_rk = dot_dict_rk[dot_index][3]
        if x_cord_rk > 699:
            len_rk +=1
        if dot_rkx < 0 or dot_rky < 0 or dot_rkx > screen_width or dot_rky > screen_height:
            dots_to_remove_rk.append(dot_index)
    for index in dots_to_remove_rk:
        del dot_dict_rk[index]
    if random.random() < expectancy_l:
        dot_dict_rl[index_rl] = (699, 499, 1.57, 583.5, 800)
        index_rl += 1
    for dot_index in dot_dict_rl:
        point_x, point_y, angle, dot_rlx, dot_rly = dot_dict_rl[dot_index]
        if dot_index-1 in dot_dict_rl:
            point_x_nxt, point_y_nxt, angle_nxt, dot_rlx_nxt, dot_rly_nxt = dot_dict_rl[dot_index - 1]
        else:
            dot_rlx_nxt, dot_rly_nxt = 0, 0
        angle, dot_rlx, dot_rly = right_turn.right_turn_l(point_x, point_y, angle, dot_rlx, dot_rly, dot_rlx_nxt, dot_rly_nxt)
        dot_dict_rl[dot_index] = (point_x, point_y, angle, dot_rlx, dot_rly)
        y_cord_rl = dot_dict_rl[dot_index][4]
        if y_cord_rl > 499:
            len_rl +=1
        if dot_rlx < 0 or dot_rly < 0 or dot_rlx > screen_width or dot_rly > screen_height:
            dots_to_remove_rl.append(dot_index)
    for index in dots_to_remove_rl:
        del dot_dict_rl[index]


    len_phase_i = len_li+len_ti+len_ri
    len_phase_j = len_lj+len_tj+len_rj
    len_phase_k = len_lk+len_tk+len_rk
    len_phase_l = len_ll+len_tl+len_rl

    if t == True:
        m = fps * (((len_phase_i**2) * 0.01)+2)
        n = fps * (((len_phase_j**2) * 0.01)+2)
        o = fps * (((len_phase_k**2) * 0.01)+2)
        s = fps * (((len_phase_l**2) * 0.01)+2)
        t = False

    if timer <= 0:
        phase_i()
    elif timer < m:
        phase_i()
    elif timer <= m+n:
        phase_j()
    elif timer <= m+n+o:
        phase_k()
    elif timer <= m+n+o+s:
        phase_l()
    else:
        t = True
        timer=0

    if karan_johar >= 60:
        density_i.append(len_phase_i)
        density_j.append(len_phase_j)
        density_k.append(len_phase_k)
        density_l.append(len_phase_l)
        karan_johar == 0
        shararti_molana += 1
        average_a = sum(density_i) / len(density_i)
        average_b = sum(density_j) / len(density_j)
        average_c = sum(density_k) / len(density_k)
        average_d = sum(density_l) / len(density_l)
        print("The average of the list is:", average_a, average_b, average_c, average_d)


    pygame.display.flip()

    clock.tick(fps)
    karan_johar += 1
    shararti_molana += 1
    frame_count += 1
    timer += 1
pygame.quit()



filename = 'Density_data.xlsx'
if not os.path.exists(filename):
    book = openpyxl.Workbook()
    book.save(filename)
else:
    book = openpyxl.load_workbook(filename)




sheet1 = book.get_sheet_by_name("semi-actuated") if "semi-actuated" in book.sheetnames else book.create_sheet("semi-actuated")


sheet1.cell(row=2, column=2, value='Density i')
sheet1.cell(row=2, column=3, value='Density j')
sheet1.cell(row=2, column=4, value='Density k')
sheet1.cell(row=2, column=5, value='Density l')
sheet1.cell(row=3, column=6, value=average_a)
sheet1.cell(row=3, column=7, value=average_b)
sheet1.cell(row=3, column=8, value=average_c)
sheet1.cell(row=3, column=9, value=average_d)


for i in range(len(density_i)):
    sheet1.cell(row=i+3, column=2, value=density_i[i])
    sheet1.cell(row=i+3, column=3, value=density_j[i])
    sheet1.cell(row=i+3, column=4, value=density_k[i])
    sheet1.cell(row=i+3, column=5, value=density_l[i])

book.save(filename)

print("Operation Successful \n Check the excel file in the current working directory")